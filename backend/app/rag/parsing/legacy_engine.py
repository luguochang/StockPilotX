from __future__ import annotations

import io
import re
import time
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Any

from .models import ParseQuality, ParseResult, ParseTrace

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - optional runtime dependency
    load_workbook = None  # type: ignore[assignment]


class LegacyParsingEngine:
    """Best-effort local parsing engine with lightweight fallbacks."""

    _TEXT_EXT = {
        ".txt",
        ".md",
        ".csv",
        ".json",
        ".log",
        ".html",
        ".htm",
        ".ts",
        ".js",
        ".py",
    }
    _IMAGE_EXT = {".png", ".jpg", ".jpeg", ".bmp", ".tiff", ".webp"}

    def supports(self, *, filename: str) -> bool:
        _ = filename
        return True

    def extract(self, *, filename: str, raw_bytes: bytes, content_type: str = "") -> ParseResult:
        started = time.perf_counter()
        ext = str(Path(filename).suffix or "").lower()
        notes: list[str] = []
        ocr_used = False

        if ext in self._TEXT_EXT:
            notes.append("plain_text_decode")
            text = self._decode_text_bytes(raw_bytes)
            return self._build_result(text=text, notes=notes, ocr_used=False, started=started)

        if ext == ".docx":
            try:
                with zipfile.ZipFile(io.BytesIO(raw_bytes)) as zf:
                    xml_bytes = zf.read("word/document.xml")
                root = ET.fromstring(xml_bytes)
                text = " ".join(x.strip() for x in root.itertext() if str(x).strip())
                notes.append("docx_xml_extract")
                return self._build_result(text=text, notes=notes, ocr_used=False, started=started)
            except Exception:  # noqa: BLE001
                notes.append("docx_extract_failed_fallback_decode")
                text = self._decode_text_bytes(raw_bytes)
                return self._build_result(text=text, notes=notes, ocr_used=False, started=started)

        if ext == ".pptx":
            try:
                with zipfile.ZipFile(io.BytesIO(raw_bytes)) as zf:
                    slide_xml_files = [
                        name
                        for name in zf.namelist()
                        if name.startswith("ppt/slides/slide") and name.endswith(".xml")
                    ]
                    parts: list[str] = []
                    for name in sorted(slide_xml_files):
                        xml_bytes = zf.read(name)
                        root = ET.fromstring(xml_bytes)
                        slide_text = " ".join(x.strip() for x in root.itertext() if str(x).strip())
                        if slide_text:
                            parts.append(slide_text)
                text = "\n".join(parts)
                notes.append("pptx_xml_extract")
                return self._build_result(text=text, notes=notes, ocr_used=False, started=started)
            except Exception:  # noqa: BLE001
                notes.append("pptx_extract_failed_fallback_decode")
                text = self._decode_text_bytes(raw_bytes)
                return self._build_result(text=text, notes=notes, ocr_used=False, started=started)

        if ext in {".xlsx", ".xlsm"} and load_workbook is not None:
            try:
                wb = load_workbook(filename=io.BytesIO(raw_bytes), read_only=True, data_only=True)
                lines: list[str] = []
                for ws in wb.worksheets[:8]:
                    lines.append(f"[sheet:{ws.title}]")
                    max_rows = 1500
                    max_cols = 32
                    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                        if row_idx > max_rows:
                            break
                        cells = [str(c).strip() for c in row[:max_cols] if c is not None and str(c).strip()]
                        if cells:
                            lines.append(" | ".join(cells))
                text = "\n".join(lines)
                notes.append("xlsx_extract")
                return self._build_result(text=text, notes=notes, ocr_used=False, started=started)
            except Exception:  # noqa: BLE001
                notes.append("xlsx_extract_failed_fallback_decode")
                text = self._decode_text_bytes(raw_bytes)
                return self._build_result(text=text, notes=notes, ocr_used=False, started=started)

        if ext == ".pdf":
            try:
                import pypdf  # type: ignore
            except Exception:  # noqa: BLE001
                notes.append("pdf_parser_unavailable")
            else:
                try:
                    reader = pypdf.PdfReader(io.BytesIO(raw_bytes))
                    pages = [str(page.extract_text() or "") for page in reader.pages]
                    text = "\n".join(x for x in pages if x.strip())
                    if text.strip():
                        notes.append("pdf_pypdf_extract")
                        return self._build_result(text=text, notes=notes, ocr_used=False, started=started)
                    notes.append("pdf_parse_failed")
                except Exception:  # noqa: BLE001
                    # Parse failures should not be mislabeled as dependency missing.
                    notes.append("pdf_parse_failed")
            ascii_chunks = re.findall(rb"[A-Za-z0-9][A-Za-z0-9 ,.;:%()\-_/]{16,}", raw_bytes)
            decoded = " ".join(x.decode("latin1", errors="ignore") for x in ascii_chunks)
            notes.append("pdf_ascii_fallback")
            if self._looks_like_pdf_binary_stream(decoded):
                notes.append("pdf_binary_stream_detected")
                decoded = ""
            return self._build_result(text=decoded, notes=notes, ocr_used=False, started=started)

        if ext in self._IMAGE_EXT:
            ocr_used = True
            try:
                from PIL import Image  # type: ignore
                import pytesseract  # type: ignore

                image = Image.open(io.BytesIO(raw_bytes))
                text = str(pytesseract.image_to_string(image, lang="chi_sim+eng") or "")
                notes.append("image_pytesseract_extract")
                return self._build_result(text=text, notes=notes, ocr_used=ocr_used, started=started)
            except Exception:  # noqa: BLE001
                notes.append("image_ocr_unavailable")
                return self._build_result(text="", notes=notes, ocr_used=ocr_used, started=started)

        notes.append(f"generic_decode:{content_type or 'unknown'}")
        text = self._decode_text_bytes(raw_bytes)
        return self._build_result(text=text, notes=notes, ocr_used=ocr_used, started=started)

    @staticmethod
    def _decode_text_bytes(raw_bytes: bytes) -> str:
        for enc in ("utf-8", "gbk", "utf-16", "latin1"):
            try:
                return raw_bytes.decode(enc, errors="ignore")
            except Exception:  # noqa: BLE001
                continue
        return ""

    @staticmethod
    def _looks_like_pdf_binary_stream(text: str) -> bool:
        normalized = str(text or "")
        if not normalized:
            return False
        markers = [
            "Filter/FlateDecode",
            "endstream",
            "stream",
            "/Length",
            "Subtype/Type1C",
            "obj",
            "endobj",
        ]
        hit_count = sum(normalized.count(marker) for marker in markers)
        return hit_count >= 8 or (hit_count >= 4 and len(normalized) >= 300)

    def _build_result(self, *, text: str, notes: list[str], ocr_used: bool, started: float) -> ParseResult:
        duration_ms = int((time.perf_counter() - started) * 1000)
        normalized = re.sub(r"\s+", " ", str(text or "")).strip()
        coverage = min(1.0, len(normalized) / 3000.0)
        garbled_count = sum(1 for ch in normalized if ch in {"?", "\ufffd"})
        garbled_ratio = (garbled_count / max(1, len(normalized))) if normalized else 0.0
        ocr_conf = 0.75 if ocr_used and normalized else (0.0 if ocr_used else 1.0)
        quality_score = max(0.0, min(1.0, coverage * 0.65 + (1.0 - garbled_ratio) * 0.35))
        if ocr_used and not normalized:
            quality_score = 0.0
        if "pdf_binary_stream_detected" in notes:
            quality_score = 0.0
        trace = ParseTrace(
            parser_name="legacy_parser",
            parser_version="1",
            ocr_used=ocr_used,
            duration_ms=duration_ms,
            notes=list(notes),
        )
        quality = ParseQuality(
            text_coverage_ratio=coverage,
            garbled_ratio=garbled_ratio,
            ocr_confidence_avg=ocr_conf,
            quality_score=quality_score,
        )
        return ParseResult(
            plain_text=str(text or ""),
            parse_note=",".join(notes),
            trace=trace,
            quality=quality,
        )
